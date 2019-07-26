import pandas as pd
import csv
import os
import os, zipfile
import openpyxl as pyxl
import logging  # for logging


def get_val_by_key(key_value, work_sheet, col_count, row_count):
    found = False
    max_search_row = 120
    max_search_col = 15
    for i in range(1, max_search_row):
        for j in range(1, max_search_col):
            cell_value = str(work_sheet.cell(row=i, column=j).value)
            if cell_value is not None:
                if key_value.lower() in cell_value.lower() and "margin" not in cell_value.lower():
                    return work_sheet.cell(row=i + row_count, column=j + col_count).value
                    found = True
                    break
        if found:
            break
    if not found:
        return None


def fill_data(sheet_name, excel_file_ob, key_value, col_count=0, row_count=0):
    if sheet_name in excel_file_ob:
        ws = excel_file_ob[sheet_name]
        return get_val_by_key(key_value, ws, col_count, row_count)
    else:
        return None


def create_row(fieldnames):
    result_dict = {}
    for key in fieldnames:
        result_dict[key] = None
    return result_dict


def parse_file(file, folder_name, fieldnames):
    result_dict = create_row(fieldnames)
    if (file[-4:] == 'xlsx'):
        print("Processing file {}".format(folder_name + '/' + file))
        xlsx_file = pyxl.load_workbook(folder_name + '/' + file, data_only=True, read_only=True)
        result_dict['column_header1'] = fill_data("Customer Details", xlsx_file, "column_header1_required_values",
                                                  row_count=1)
        result_dict['column_header2'] = fill_data("Customer Details", xlsx_file, "column_header2_required_values",
                                                  row_count=1)
        result_dict['column_header3'] = fill_data("Customer Details", xlsx_file, "column_header3_required_values",
                                                  row_count=2)
        result_dict['column_header4'] = fill_data("Customer Details", xlsx_file, "column_header4_required_values",
                                                  row_count=3)
        result_dict['column_header5'] = fill_data("Customer Details", xlsx_file, "column_header5_required_values",
                                                  col_count=2)
        result_dict['column_header6'] = fill_data("Customer Details", xlsx_file, "column_header6_required_values",
                                                  col_count=2)
        result_dict['column_header7'] = fill_data("Customer Details", xlsx_file, "column_header7_required_values",
                                                  row_count=1)
        result_dict['column_header8'] = fill_data("Customer Details", xlsx_file, "column_header8_required_values",
                                                  row_count=1)
        result_dict['column_header9'] = fill_data("Customer Details", xlsx_file, "column_header9_required_values",
                                                  col_count=2)
        result_dict['column_header10'] = fill_data("Customer Details", xlsx_file, "column_header10_required_values",
                                                   col_count=2)
        result_dict['column_header11'] = fill_data("Required_sheet_name", xlsx_file, "column_header11_required_values",
                                                   col_count=1)
        result_dict['column_header12'] = fill_data("Required_sheet_name", xlsx_file, "column_header12_required_values",
                                                   col_count=1)
        result_dict['column_header13'] = fill_data("Required_sheet_name", xlsx_file, "column_header13_required_values",
                                                   col_count=3)
    return result_dict


def check_error(logger, result_dict, filename):
    for key in result_dict:
        if result_dict[key] is None:
            logger.error("{} - {} not present".format(filename, key))


if __name__ == '__main__':
    xlsx_data_dir = 'required path containing sheets from where data is extracted'
    output_dump_path = "path to csv to which we have to dump data.csv"
    # logging config
    logging.basicConfig(filename='cet_parser.log', level=logging.INFO)
    logger = logging.getLogger()

    # initialise dict array
    dict_array = []
    fieldnames = ['Name', 'Id', 'Score', 'Receipts', 'lev', 'Programme',
                  'dname', 'sname', 'con', 'ind', 'header1', 'header2', 'header3'] - ---> These
    will
    be
    the
    headers in the
    file
    under
    which
    data
    will
    be
    printed.

    # zip coversion
    extension = ".zip"
    for path, dir_list, file_list in os.walk(xlsx_data_dir):
        for file_name in file_list:
            if file_name.endswith(".zip"):
                abs_file_path = os.path.join(path, file_name)
                parent_path = os.path.split(abs_file_path)[0]
                output_folder_name = os.path.splitext(abs_file_path)[0]
                output_path = os.path.join(parent_path, output_folder_name)
                zip_obj = zipfile.ZipFile(abs_file_path, 'r')
                zip_obj.extractall(output_path)
                zip_obj.close()

    for folder_name, _, file_list in os.walk(xlsx_data_dir):
        for file in file_list:
            try:
                logger.info('Process started for file {}'.format(file))
                result_dict = parse_file(file, folder_name, fieldnames)
                check_error(logger, result_dict, file)
                dict_array.append(result_dict)
            except Exception as e:
                raise e
                logger.error("Process stopped for file {} with error {}".format(file, str(e)))

    with open(output_dump_path, 'w', newline='') as output_file:
        dict_writer = csv.DictWriter(output_file, fieldnames)
        dict_writer.writeheader()
        dict_writer.writerows(dict_array)
    print('HURRAY!!!! All converted')




